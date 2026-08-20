from openpyxl import load_workbook
from django.db import transaction

from .models import Question, Choice


def import_questions_from_xlsx(exam, excel_file):
    """
    Import objective questions from an Excel (.xlsx) file into the given exam.

    All worksheets are processed and their questions are combined into
    one exam.

    Expected columns (row 1 = header):
        Question | Option A | Option B | Option C | Option D | Correct Answer | Marks

    Rules:
        - All worksheets are imported.
        - Correct Answer must be A, B, C or D (case-insensitive).
        - All four options are required.
        - Blank rows are skipped.
        - Duplicate question text (already present on this exam) is ignored.
        - Marks default to 1 if missing or invalid.

    Returns:
        dict with keys:
        imported (int), skipped_duplicates (int), errors (list of str)
    """

    result = {
        'imported': 0,
        'skipped_duplicates': 0,
        'errors': [],
    }

    try:
        wb = load_workbook(excel_file, data_only=True)
    except Exception:
        result['errors'].append(
            'Could not read the Excel file. Please check the format.'
        )
        return result

    # Existing question texts for this exam
    # Used for duplicate detection across ALL sheets.
    existing_texts = set(
        t.strip().lower()
        for t in exam.questions.values_list('text', flat=True)
        if t
    )

    with transaction.atomic():

        # Process EVERY worksheet in the workbook
        for ws in wb.worksheets:

            for row_idx, row in enumerate(
                ws.iter_rows(min_row=2, values_only=True),
                start=2
            ):

                # Skip completely blank rows
                if not row or all(
                    cell is None or str(cell).strip() == ''
                    for cell in row
                ):
                    continue

                try:
                    question_text = (
                        str(row[0]).strip()
                        if row[0] is not None
                        else ''
                    )

                    opt_a = (
                        str(row[1]).strip()
                        if len(row) > 1 and row[1] is not None
                        else ''
                    )

                    opt_b = (
                        str(row[2]).strip()
                        if len(row) > 2 and row[2] is not None
                        else ''
                    )

                    opt_c = (
                        str(row[3]).strip()
                        if len(row) > 3 and row[3] is not None
                        else ''
                    )

                    opt_d = (
                        str(row[4]).strip()
                        if len(row) > 4 and row[4] is not None
                        else ''
                    )

                    correct = (
                        str(row[5]).strip().upper()
                        if len(row) > 5 and row[5] is not None
                        else ''
                    )

                    marks_raw = row[6] if len(row) > 6 else 1

                except Exception:
                    result['errors'].append(
                        f'Sheet "{ws.title}", row {row_idx}: '
                        'could not read data.'
                    )
                    continue

                if not question_text:
                    result['errors'].append(
                        f'Sheet "{ws.title}", row {row_idx}: '
                        'Question is empty.'
                    )
                    continue

                if not all([opt_a, opt_b, opt_c, opt_d]):
                    result['errors'].append(
                        f'Sheet "{ws.title}", row {row_idx}: '
                        'All four options (A–D) are required.'
                    )
                    continue

                if correct not in ('A', 'B', 'C', 'D'):
                    result['errors'].append(
                        f'Sheet "{ws.title}", row {row_idx}: '
                        f"Correct Answer must be A, B, C or D "
                        f"(got '{correct}')."
                    )
                    continue

                # Duplicate check across the entire exam
                question_key = question_text.lower()

                if question_key in existing_texts:
                    result['skipped_duplicates'] += 1
                    continue

                # Parse marks
                try:
                    marks = int(marks_raw) if marks_raw is not None else 1

                    if marks < 1:
                        marks = 1

                except (ValueError, TypeError):
                    marks = 1

                # Create question
                question = Question.objects.create(
                    exam=exam,
                    text=question_text,
                    marks=marks,
                )

                # Create choices
                choices_map = {
                    'A': opt_a,
                    'B': opt_b,
                    'C': opt_c,
                    'D': opt_d,
                }

                for letter, text in choices_map.items():
                    Choice.objects.create(
                        question=question,
                        text=text,
                        is_correct=(letter == correct),
                    )

                existing_texts.add(question_key)
                result['imported'] += 1

    return result